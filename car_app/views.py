from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum
from datetime import datetime, timedelta
from .models import Car, Sale, MonthlyExpense
from .forms import CarForm, SaleForm, MonthlyExpenseForm
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
import json
import os
import arabic_reshaper
from bidi.algorithm import get_display
from .font_manager import register_arabic_fonts, get_arabic_font_name, get_arabic_font_bold
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# تسجيل الخطوط العربية عند بدء التطبيق
register_arabic_fonts()

def reshape_arabic_text(text):
    """Reshape Arabic text for proper display in PDF"""
    if not text:
        return text
    try:
        # تطبيق arabic_reshaper و bidi
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except Exception as e:
        # إذا فشل، أرجع النص الأصلي
        return text



@login_required
def dashboard(request):
    cars = Car.objects.filter(user=request.user)
    sales = Sale.objects.filter(car__user=request.user)
    expenses = MonthlyExpense.objects.filter(user=request.user)
    
    # Filter by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        sales = sales.filter(sale_date__range=[start_date, end_date])
        expenses = expenses.filter(date__range=[start_date, end_date])
    
    # Calculate totals
    total_partial_profit = sales.aggregate(Sum('partial_profit'))['partial_profit__sum'] or 0
    # Try DB aggregation first, fallback to Python sum if necessary
    from decimal import Decimal
    agg = expenses.aggregate(total=Sum('amount'))['total']
    if agg is None:
        if expenses.exists():
            total_expenses = sum(e.amount for e in expenses)
        else:
            total_expenses = Decimal('0.00')
    else:
        total_expenses = agg
    net_profit = total_partial_profit - total_expenses
    
    total_cars = cars.count()
    sold_cars = cars.filter(status='sold').count()
    available_cars = cars.filter(status='available').count()
    
    context = {
        'total_cars': total_cars,
        'sold_cars': sold_cars,
        'available_cars': available_cars,
        'total_partial_profit': total_partial_profit,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'start_date': start_date or '',
        'end_date': end_date or '',
    }
    
    return render(request, 'car_app/dashboard.html', context)


@login_required
def car_list(request):
    cars = Car.objects.filter(user=request.user)
    
    # Search
    search_query = request.GET.get('search')
    if search_query:
        cars = cars.filter(Q(car_type__icontains=search_query) | 
                          Q(chassis_number__icontains=search_query))
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        cars = cars.filter(status=status_filter)
    
    # Filter by clearance type (purchase/auction)
    clearance_filter = request.GET.get('clearance')
    if clearance_filter:
        cars = cars.filter(clearance_type=clearance_filter)
    
    # Filter by month/year (purchase_date)
    month_filter = request.GET.get('month')
    year_filter = request.GET.get('year')
    try:
        if month_filter:
            month_int = int(month_filter)
            cars = cars.filter(purchase_date__month=month_int)
        if year_filter:
            year_int = int(year_filter)
            cars = cars.filter(purchase_date__year=year_int)
    except ValueError:
        month_filter = None
        year_filter = None
    
    context = {
        'cars': cars,
        'search_query': search_query or '',
        'status_filter': status_filter or '',
        'clearance_filter': clearance_filter or '',
        'month_filter': month_filter or '',
        'year_filter': year_filter or '',
    }
    
    return render(request, 'car_app/car_list.html', context)


@login_required
def export_cars_pdf(request):
    """Export cars list to PDF using ReportLab with Arabic font support"""
    cars = Car.objects.filter(user=request.user)
    
    # Apply same filters as car_list
    search_query = request.GET.get('search')
    if search_query:
        cars = cars.filter(Q(car_type__icontains=search_query) | 
                          Q(chassis_number__icontains=search_query))
    
    status_filter = request.GET.get('status')
    if status_filter:
        cars = cars.filter(status=status_filter)
    
    # Apply clearance filter if present
    clearance_filter = request.GET.get('clearance')
    if clearance_filter:
        cars = cars.filter(clearance_type=clearance_filter)
    
    # Filter by month/year (purchase_date)
    month = request.GET.get('month')
    year = request.GET.get('year')
    try:
        if month:
            month_int = int(month)
            cars = cars.filter(purchase_date__month=month_int)
        if year:
            year_int = int(year)
            cars = cars.filter(purchase_date__year=year_int)
    except ValueError:
        # ignore invalid month/year values
        pass
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="cars_list.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.75*inch, bottomMargin=0.75*inch)
    elements = []
    
    # استخدم الخطوط المسجلة بشكل صحيح
    title_font = get_arabic_font_bold()
    body_font = get_arabic_font_name()
    
    # Title style
    title_style = ParagraphStyle(
        'Title',
        parent=getSampleStyleSheet()['Heading1'],
        alignment=TA_CENTER,
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        fontName=title_font,
        spaceAfter=20,
    )
    
    # Table data with reshaped Arabic text
    headers = [
        reshape_arabic_text('الاسم'),
        reshape_arabic_text('النوع'),
        reshape_arabic_text('السنة'),
        reshape_arabic_text('الشاصي'),
        reshape_arabic_text('تاريخ الشراء'),
        reshape_arabic_text('القيمة'),
        reshape_arabic_text('التخليص'),
        reshape_arabic_text('الحالة')
    ]
    data = [headers]
    
    for car in cars:
        status_text = reshape_arabic_text('مباع' if car.status == 'sold' else 'متاح')
        clearance_text = reshape_arabic_text('شراء' if car.clearance_type == 'purchase' else 'اعلان')
        car_type = car.get_car_type_display()
        
        row = [
            reshape_arabic_text(car.name or 'N/A'),
            reshape_arabic_text(car_type),
            str(car.year),
            car.chassis_number,
            car.purchase_date.strftime('%d/%m/%Y'),
            f'{car.purchase_value:.2f}',
            clearance_text,
            status_text,
        ]
        data.append(row)
    
    # Add title
    title = Paragraph(reshape_arabic_text('قائمة السيارات'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Create table with proper styling
    table = Table(data, colWidths=[1.2*inch, 1.2*inch, 0.7*inch, 0.9*inch, 1*inch, 0.8*inch, 0.7*inch, 0.7*inch])
    
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), title_font),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Body styling
        ('FONTNAME', (0, 1), (-1, -1), body_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    
    elements.append(table)
    
    # Build PDF
    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f'خطأ في إنشاء PDF: {str(e)}', status=400)
    
    return response


@login_required
def add_car(request):
    if request.method == 'POST':
        form = CarForm(request.POST)
        if form.is_valid():
            car = form.save(commit=False)
            car.user = request.user
            car.save()
            messages.success(request, f'✓ تم إضافة السيارة "{car.name}" بنجاح!')
            return redirect('car_list')
        else:
            messages.error(request, '❌ حدث خطأ في النموذج. تحقق من البيانات المدخلة.')
    else:
        form = CarForm()
    
    return render(request, 'car_app/add_car.html', {'form': form})


@login_required
def edit_car(request, car_id):
    car = get_object_or_404(Car, id=car_id, user=request.user)
    
    if request.method == 'POST':
        form = CarForm(request.POST, instance=car)
        if form.is_valid():
            form.save()
            messages.success(request, f'✓ تم حفظ تعديلات السيارة "{car.name}" بنجاح!')
            return redirect('car_list')
        else:
            messages.error(request, '❌ حدث خطأ في النموذج. تحقق من البيانات المدخلة.')
    else:
        form = CarForm(instance=car)
    
    return render(request, 'car_app/edit_car.html', {'form': form, 'car': car})


@login_required
@require_POST
def delete_car(request, car_id):
    car = get_object_or_404(Car, id=car_id, user=request.user)
    car.delete()
    return redirect('car_list')


@login_required
@login_required
def add_sale(request, car_id):
    car = get_object_or_404(Car, id=car_id, user=request.user)
    
    if request.method == 'POST':
        form = SaleForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)
            sale.car = car
            sale.save()
            car.status = 'sold'
            car.save()
            messages.success(request, f'✓ تم تسجيل بيع السيارة "{car.name}" بنجاح!')
            return redirect('sales_list')
        else:
            messages.error(request, '❌ حدث خطأ في النموذج. تحقق من البيانات المدخلة.')
    else:
        form = SaleForm()
    
    context = {
        'form': form,
        'car': car,
    }
    return render(request, 'car_app/add_sale.html', context)


@login_required
def sales_list(request):
    sales = Sale.objects.filter(car__user=request.user)

    # Date range filter (start_date and end_date expected in YYYY-MM-DD)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        try:
            sd = datetime.strptime(start_date, '%Y-%m-%d').date()
            ed = datetime.strptime(end_date, '%Y-%m-%d').date()
            sales = sales.filter(sale_date__range=[sd, ed])
        except ValueError:
            # ignore invalid date formats
            pass

    context = {
        'sales': sales,
        'start_date': start_date or '',
        'end_date': end_date or '',
    }

    return render(request, 'car_app/sales_list.html', context)


@login_required
def export_sales_pdf(request):
    """Export sales list to PDF with car details - compatible with PythonAnywhere"""
    sales = Sale.objects.filter(car__user=request.user)

    # Date range filter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        try:
            sd = datetime.strptime(start_date, '%Y-%m-%d').date()
            ed = datetime.strptime(end_date, '%Y-%m-%d').date()
            sales = sales.filter(sale_date__range=[sd, ed])
        except ValueError:
            pass

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_list.pdf"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.75*inch, bottomMargin=0.75*inch)
    elements = []

    # استخدم الخطوط المسجلة بشكل صحيح
    title_font = get_arabic_font_bold()
    body_font = get_arabic_font_name()

    # Title style
    title_style = ParagraphStyle(
        'Title',
        parent=getSampleStyleSheet()['Heading1'],
        alignment=TA_CENTER,
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        fontName=title_font,
        spaceAfter=20,
    )

    # Table headers
    headers = [
        reshape_arabic_text('اسم السيارة'),
        reshape_arabic_text('النوع'),
        reshape_arabic_text('السنة'),
        reshape_arabic_text('الشاصي'),
        reshape_arabic_text('تاريخ الشراء'),
        reshape_arabic_text('قيمة الشراء'),
        reshape_arabic_text('التخليص')
    ]
    data = [headers]

    # Add sales data
    for sale in sales:
        car = sale.car
        row = [
            reshape_arabic_text(car.name or 'N/A'),
            reshape_arabic_text(car.get_car_type_display()),
            str(car.year),
            car.chassis_number,
            car.purchase_date.strftime('%d/%m/%Y'),
            f'{car.purchase_value:.2f}',
            reshape_arabic_text(car.get_clearance_type_display()),
        ]
        data.append(row)

    # Add title
    title = Paragraph(reshape_arabic_text('قائمة المبيعات'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))

    # Create table
    table = Table(data, colWidths=[1.2*inch, 1*inch, 0.7*inch, 0.9*inch, 1*inch, 0.8*inch, 0.9*inch])

    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), title_font),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Body styling
        ('FONTNAME', (0, 1), (-1, -1), body_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))

    elements.append(table)

    # Build PDF
    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f'خطأ في إنشاء PDF: {str(e)}', status=400)

    return response


@login_required
def export_sales_excel(request):
    """Export sales list to Excel with car details"""
    from openpyxl.utils import get_column_letter
    
    sales = Sale.objects.filter(car__user=request.user)

    # Date range filter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        try:
            sd = datetime.strptime(start_date, '%Y-%m-%d').date()
            ed = datetime.strptime(end_date, '%Y-%m-%d').date()
            sales = sales.filter(sale_date__range=[sd, ed])
        except ValueError:
            pass

    # Create Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "المبيعات"

    # Headers
    headers = [
        'اسم السيارة',
        'النوع',
        'السنة',
        'الشاصي',
        'تاريخ الشراء',
        'قيمة الشراء',
        'التخليص'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, name='Arial', size=11)
        cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Set column widths
    column_widths = [20, 15, 10, 15, 15, 15, 15]
    for col_num, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = width

    # Add borders style
    thin_border = Border(
        left=Side(style='thin', color='bdc3c7'),
        right=Side(style='thin', color='bdc3c7'),
        top=Side(style='thin', color='bdc3c7'),
        bottom=Side(style='thin', color='bdc3c7')
    )

    # Add data rows
    for row_num, sale in enumerate(sales, 2):
        car = sale.car
        row_data = [
            car.name or 'N/A',
            car.get_car_type_display(),
            str(car.year),
            car.chassis_number,
            car.purchase_date.strftime('%d/%m/%Y'),
            f'{car.purchase_value:.2f}',
            car.get_clearance_type_display(),
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='f9f9f9', end_color='f9f9f9', fill_type='solid')

    # Freeze header row
    worksheet.freeze_panes = 'A2'

    # Add header borders
    for col_num in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.border = thin_border

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales_list.xlsx"'
    workbook.save(response)
    return response


@login_required
def edit_sale(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id, car__user=request.user)
    
    if request.method == 'POST':
        form = SaleForm(request.POST, instance=sale)
        if form.is_valid():
            form.save()
            messages.success(request, '✓ تم حفظ تعديلات البيع بنجاح!')
            return redirect('sales_list')
        else:
            messages.error(request, '❌ حدث خطأ في النموذج. تحقق من البيانات المدخلة.')
    else:
        form = SaleForm(instance=sale)
    
    return render(request, 'car_app/edit_sale.html', {'form': form, 'sale': sale})


@login_required
def expenses_list(request):
    expenses = MonthlyExpense.objects.filter(user=request.user)

    # Calculate total expenses (DB aggregate with Python fallback)
    from decimal import Decimal
    agg = expenses.aggregate(total=Sum('amount'))['total']
    if agg is None:
        if expenses.exists():
            total_expenses = sum(e.amount for e in expenses)
        else:
            total_expenses = Decimal('0.00')
    else:
        total_expenses = agg

    context = {
        'expenses': expenses,
        'total_expenses': total_expenses,
    }

    return render(request, 'car_app/expenses_list.html', context)


@login_required
@login_required
def add_expense(request):
    if request.method == 'POST':
        form = MonthlyExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            messages.success(request, f'✓ تم إضافة المصروف "{expense.description}" بنجاح!')
            return redirect('expenses_list')
        else:
            messages.error(request, '❌ حدث خطأ في النموذج. تحقق من البيانات المدخلة.')
    else:
        form = MonthlyExpenseForm()
    
    return render(request, 'car_app/add_expense.html', {'form': form})


@login_required
@require_POST
def delete_expense(request, expense_id):
    expense = get_object_or_404(MonthlyExpense, id=expense_id, user=request.user)
    expense.delete()
    return redirect('expenses_list')


# API Views
@login_required
def api_car_list(request):
    cars = Car.objects.filter(user=request.user)
    
    search_query = request.GET.get('search')
    if search_query:
        cars = cars.filter(Q(car_type__icontains=search_query) | 
                          Q(chassis_number__icontains=search_query))
    
    status_filter = request.GET.get('status')
    if status_filter:
        cars = cars.filter(status=status_filter)
    
    data = [
        {
            'id': car.id,
            'car_type': car.get_car_type_display(),
            'year': car.year,
            'chassis_number': car.chassis_number,
            'purchase_date': car.purchase_date.isoformat(),
            'purchase_value': float(car.purchase_value),
            'status': car.get_status_display(),
        }
        for car in cars
    ]
    
    return JsonResponse(data, safe=False)


@login_required
def api_sales_list(request):
    sales = Sale.objects.filter(car__user=request.user)
    
    data = [
        {
            'id': sale.id,
            'car': sale.car.get_car_type_display(),
            'chassis_number': sale.car.chassis_number,
            'sale_date': sale.sale_date.isoformat(),
            'sale_value': float(sale.sale_value),
            'purchase_value': float(sale.car.purchase_value),
            'total_profit': float(sale.get_total_profit()),
            'partial_profit': float(sale.partial_profit),
        }
        for sale in sales
    ]
    
    return JsonResponse(data, safe=False)
